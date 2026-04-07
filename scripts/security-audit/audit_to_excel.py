#!/usr/bin/env python3
"""보안 점검 JSON 결과를 Excel 보고서로 변환하는 스크립트.

사용법:
    python audit_to_excel.py <json파일경로> [출력엑셀경로]
"""
import json
import re as _re
import sys
import os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.chart import PieChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl 패키지가 필요합니다. 설치: pip install openpyxl")
    sys.exit(1)


# ── 색상/스타일 상수 ──────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(name="맑은 고딕", bold=True, size=14)
SUB_FONT    = Font(name="맑은 고딕", bold=True, size=11)
NORMAL_FONT = Font(name="맑은 고딕", size=10)
WRAP_ALIGN  = Alignment(wrap_text=True, vertical="center", horizontal="left")
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

VERDICT_COLORS = {
    "양호":       PatternFill("solid", fgColor="C6EFCE"),
    "취약":       PatternFill("solid", fgColor="FFC7CE"),
    "N/A":        PatternFill("solid", fgColor="D9D9D9"),
    "수동점검필요": PatternFill("solid", fgColor="FFEB9C"),
}
VERDICT_FONTS = {
    "양호":       Font(name="맑은 고딕", size=10, color="006100"),
    "취약":       Font(name="맑은 고딕", size=10, color="9C0006"),
    "N/A":        Font(name="맑은 고딕", size=10, color="404040"),
    "수동점검필요": Font(name="맑은 고딕", size=10, color="9C6500"),
}
IMPORTANCE_FILL = {
    "상": PatternFill("solid", fgColor="FCE4EC"),
    "중": PatternFill("solid", fgColor="FFF9C4"),
    "하": PatternFill("solid", fgColor="E8F5E9"),
}

# ── 카테고리 매핑 ─────────────────────────────────────────────────
UNIX_CATEGORY_MAP = {
    range(1, 14):  "계정관리",
    range(14, 34): "파일 및 디렉터리 관리",
    range(34, 62): "서비스 관리",
    range(62, 65): "패치 관리",
    range(65, 68): "로그 관리",
}

WINDOWS_CATEGORY_MAP = {
    range(1, 15):  "계정 관리",
    range(15, 38): "서비스 관리",
    range(38, 40): "패치 관리",
    range(40, 44): "로그 관리",
    range(44, 63): "보안 관리",
}

def get_category(code: str) -> str:
    import re
    m = re.match(r'^([A-Za-z]+)-?(\d+)$', code.strip())
    if not m:
        return "기타"
    prefix = m.group(1).upper()
    num = int(m.group(2))
    cat_map = WINDOWS_CATEGORY_MAP if prefix == "W" else UNIX_CATEGORY_MAP
    for rng, cat in cat_map.items():
        if num in rng:
            return cat
    return "기타"


def apply_row_border(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER


def evidence_to_str(ev) -> str:
    """evidence dict/list를 읽기 쉬운 문자열로 변환."""
    if isinstance(ev, dict):
        lines = []
        for k, v in ev.items():
            if isinstance(v, list):
                if len(v) <= 5:
                    val = ", ".join(str(x) for x in v)
                else:
                    val = ", ".join(str(x) for x in v[:5]) + f" ... 외 {len(v)-5}건"
            elif isinstance(v, dict):
                val = "; ".join(f"{sk}={sv}" for sk, sv in v.items())
            else:
                val = str(v)
            lines.append(f"[{k}] {val}")
        return _sanitize("\n".join(lines))
    if isinstance(ev, list):
        return _sanitize(", ".join(str(x) for x in ev))
    return _sanitize(str(ev))


# openpyxl 불허 제어문자 제거 (탭·개행 제외)
_ILLEGAL_RE = _re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def _sanitize(s: str) -> str:
    return _ILLEGAL_RE.sub('', s)


# ═══════════════════════════════════════════════════════════════════
#  Sheet 1: 통계 및 요약
# ═══════════════════════════════════════════════════════════════════
def build_summary_sheet(wb, data):
    ws = wb.active
    ws.title = "통계 및 요약"
    ws.sheet_properties.tabColor = "2F5496"

    scan = data["scan_info"]
    summary = data["summary"]
    results = data["results"]

    # ── 플랫폼 판별 ─────────────────────────────────────────────────
    is_windows = "windows" in scan.get("os", "").lower()

    # ── 상단 타이틀 ───────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = ("Windows 서버 취약점 진단 결과 보고서"
               if is_windows
               else "Unix/Linux 서버 취약점 진단 결과 보고서")
    c.font = TITLE_FONT
    c.alignment = CENTER

    # ── 스캔 정보 ─────────────────────────────────────────────────
    if is_windows:
        info_labels = ["호스트명", "OS", "점검일시", "총 점검 항목"]
        info_values = [
            scan["hostname"], scan["os"],
            scan["scan_time"], scan["total_checks"],
        ]
    else:
        info_labels = ["호스트명", "OS", "배포판", "점검일시", "총 점검 항목"]
        info_values = [
            scan["hostname"], scan["os"], scan.get("distro", ""),
            scan["scan_time"], scan["total_checks"],
        ]
    for i, (lbl, val) in enumerate(zip(info_labels, info_values)):
        r = 3 + i
        ws.cell(row=r, column=1, value=lbl).font = SUB_FONT
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="D6E4F0")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.cell(row=r, column=2, value=str(val)).font = NORMAL_FONT
        ws.cell(row=r, column=2).alignment = CENTER
        for cc in range(1, 4):
            ws.cell(row=r, column=cc).border = THIN_BORDER

    # ── 판정 통계 테이블 ──────────────────────────────────────────
    stat_row = 9
    ws.merge_cells(f"A{stat_row}:D{stat_row}")
    ws.cell(row=stat_row, column=1, value="판정 결과 통계").font = SUB_FONT
    ws.cell(row=stat_row, column=1).alignment = CENTER

    stat_row += 1
    stat_headers = ["판정", "건수", "비율(%)"]
    for ci, h in enumerate(stat_headers, 1):
        cell = ws.cell(row=stat_row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    total = sum(summary.values()) or 1
    verdict_order = ["양호", "취약", "N/A", "수동점검필요"]
    for vi, v in enumerate(verdict_order):
        r = stat_row + 1 + vi
        cnt = summary.get(v, 0)
        ws.cell(row=r, column=1, value=v).font = VERDICT_FONTS.get(v, NORMAL_FONT)
        ws.cell(row=r, column=1).fill = VERDICT_COLORS.get(v, PatternFill())
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=2, value=cnt).font = NORMAL_FONT
        ws.cell(row=r, column=2).alignment = CENTER
        pct_cell = ws.cell(row=r, column=3, value=round(cnt / total * 100, 1))
        pct_cell.font = NORMAL_FONT
        pct_cell.alignment = CENTER
        pct_cell.number_format = "0.0"
        apply_row_border(ws, r, 1, 3)

    # ── 파이 차트 ─────────────────────────────────────────────────
    chart = PieChart()
    chart.title = "판정 분포"
    chart.width = 14
    chart.height = 10
    labels = Reference(ws, min_col=1, min_row=stat_row + 1,
                       max_row=stat_row + len(verdict_order))
    values = Reference(ws, min_col=2, min_row=stat_row,
                       max_row=stat_row + len(verdict_order))
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(labels)
    chart.style = 10
    ws.add_chart(chart, "E9")

    # ── 카테고리별 통계 ───────────────────────────────────────────
    cat_row = stat_row + len(verdict_order) + 18  # 차트 아래로 충분히 이동
    ws.merge_cells(f"A{cat_row}:F{cat_row}")
    ws.cell(row=cat_row, column=1, value="카테고리별 통계").font = SUB_FONT
    ws.cell(row=cat_row, column=1).alignment = CENTER

    cat_row += 1
    cat_headers = ["카테고리", "전체", "양호", "취약", "N/A", "수동점검"]
    for ci, h in enumerate(cat_headers, 1):
        cell = ws.cell(row=cat_row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    cat_stats = {}
    for item in results:
        cat = get_category(item["항목코드"])
        if cat not in cat_stats:
            cat_stats[cat] = {"전체": 0, "양호": 0, "취약": 0, "N/A": 0, "수동점검필요": 0}
        cat_stats[cat]["전체"] += 1
        v = item["판정"]
        if v in cat_stats[cat]:
            cat_stats[cat][v] += 1

    # 플랫폼에 따라 카테고리 순서 결정
    if is_windows:
        cat_order = ["계정 관리", "서비스 관리", "패치 관리", "로그 관리", "보안 관리"]
    else:
        cat_order = ["계정관리", "파일 및 디렉터리 관리", "서비스 관리", "패치 관리", "로그 관리"]
    for ci2, cat in enumerate(cat_order):
        r = cat_row + 1 + ci2
        s = cat_stats.get(cat, {})
        ws.cell(row=r, column=1, value=cat).font = NORMAL_FONT
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=2, value=s.get("전체", 0)).alignment = CENTER
        ws.cell(row=r, column=3, value=s.get("양호", 0)).alignment = CENTER
        ws.cell(row=r, column=4, value=s.get("취약", 0)).alignment = CENTER
        ws.cell(row=r, column=5, value=s.get("N/A", 0)).alignment = CENTER
        ws.cell(row=r, column=6, value=s.get("수동점검필요", 0)).alignment = CENTER
        for cc2 in range(1, 7):
            ws.cell(row=r, column=cc2).font = NORMAL_FONT
            ws.cell(row=r, column=cc2).border = THIN_BORDER

    # ── 전체 항목 요약 테이블 ─────────────────────────────────────
    tbl_row = cat_row + len(cat_order) + 3
    ws.merge_cells(f"A{tbl_row}:H{tbl_row}")
    ws.cell(row=tbl_row, column=1, value="전체 항목 판정 결과").font = SUB_FONT
    ws.cell(row=tbl_row, column=1).alignment = CENTER

    tbl_row += 1
    tbl_headers = ["No", "항목코드", "항목명", "카테고리", "중요도", "판정"]
    for ci3, h in enumerate(tbl_headers, 1):
        cell = ws.cell(row=tbl_row, column=ci3, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for idx, item in enumerate(results, 1):
        r = tbl_row + idx
        cat = get_category(item["항목코드"])
        verdict = item["판정"]
        imp = item.get("중요도", "")

        ws.cell(row=r, column=1, value=idx).alignment = CENTER
        ws.cell(row=r, column=2, value=item["항목코드"]).alignment = CENTER
        ws.cell(row=r, column=3, value=item["항목명"]).alignment = Alignment(vertical="center")
        ws.cell(row=r, column=4, value=cat).alignment = CENTER
        imp_cell = ws.cell(row=r, column=5, value=imp)
        imp_cell.alignment = CENTER
        if imp in IMPORTANCE_FILL:
            imp_cell.fill = IMPORTANCE_FILL[imp]
        v_cell = ws.cell(row=r, column=6, value=verdict)
        v_cell.alignment = CENTER
        v_cell.fill = VERDICT_COLORS.get(verdict, PatternFill())
        v_cell.font = VERDICT_FONTS.get(verdict, NORMAL_FONT)

        for cc3 in range(1, 7):
            ws.cell(row=r, column=cc3).border = THIN_BORDER
            if ws.cell(row=r, column=cc3).font == Font():
                ws.cell(row=r, column=cc3).font = NORMAL_FONT

    # ── 열 너비 ───────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 44
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 4
    ws.column_dimensions["H"].width = 4


# ═══════════════════════════════════════════════════════════════════
#  Sheet 2: 상세 진단
# ═══════════════════════════════════════════════════════════════════
def build_detail_sheet(wb, data):
    ws = wb.create_sheet("상세 진단")
    ws.sheet_properties.tabColor = "C00000"

    results = data["results"]

    # ── 헤더 ──────────────────────────────────────────────────────
    headers = ["No", "항목코드", "항목명", "카테고리", "중요도", "판정",
               "점검목적", "진단 근거 (Evidence)", "점검 코드 (Python)"]
    col_widths = [6, 10, 36, 20, 8, 14, 44, 80, 90]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── 데이터 행 ─────────────────────────────────────────────────
    for idx, item in enumerate(results, 1):
        r = idx + 1
        cat = get_category(item["항목코드"])
        verdict = item["판정"]
        imp = item.get("중요도", "")
        ev_str = evidence_to_str(item.get("evidence", {}))

        ws.cell(row=r, column=1, value=idx).alignment = CENTER
        ws.cell(row=r, column=2, value=item["항목코드"]).alignment = CENTER
        ws.cell(row=r, column=3, value=item["항목명"]).alignment = WRAP_ALIGN
        ws.cell(row=r, column=4, value=cat).alignment = CENTER
        imp_cell = ws.cell(row=r, column=5, value=imp)
        imp_cell.alignment = CENTER
        if imp in IMPORTANCE_FILL:
            imp_cell.fill = IMPORTANCE_FILL[imp]
        v_cell = ws.cell(row=r, column=6, value=verdict)
        v_cell.alignment = CENTER
        v_cell.fill = VERDICT_COLORS.get(verdict, PatternFill())
        v_cell.font = VERDICT_FONTS.get(verdict, NORMAL_FONT)
        ws.cell(row=r, column=7, value=item.get("점검목적", "")).alignment = WRAP_ALIGN
        ws.cell(row=r, column=8, value=ev_str).alignment = WRAP_ALIGN

        # 점검 코드
        code_str = item.get("점검코드", "")
        code_cell = ws.cell(row=r, column=9, value=code_str)
        code_cell.alignment = WRAP_ALIGN
        code_cell.font = Font(name="Consolas", size=9)

        for cc in range(1, 10):
            ws.cell(row=r, column=cc).border = THIN_BORDER
            if cc < 9 and ws.cell(row=r, column=cc).font == Font():
                ws.cell(row=r, column=cc).font = NORMAL_FONT

        # 행 높이 — evidence/코드 줄 수에 맞춰 자동 조정
        code_lines = code_str.count("\n") + 1 if code_str else 0
        line_count = max(ev_str.count("\n") + 1, code_lines, 2)
        ws.row_dimensions[r].height = min(14 * line_count, 300)

    # 필터 & 고정 틀
    ws.auto_filter.ref = f"A1:I{len(results) + 1}"
    ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════════════════════════
#  main
# ═══════════════════════════════════════════════════════════════════
def main():
    if len(sys.argv) < 2:
        print("사용법: python audit_to_excel.py <json파일> [출력엑셀파일]")
        sys.exit(1)

    json_path = sys.argv[1]
    if not os.path.isfile(json_path):
        print(f"파일을 찾을 수 없습니다: {json_path}")
        sys.exit(1)

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # 출력 파일명 결정
    if len(sys.argv) >= 3:
        out_path = sys.argv[2]
    else:
        base = os.path.splitext(os.path.basename(json_path))[0]
        out_dir = os.path.dirname(json_path) or "."
        out_path = os.path.join(out_dir, f"{base}.xlsx")

    wb = Workbook()
    build_summary_sheet(wb, data)
    build_detail_sheet(wb, data)

    wb.save(out_path)
    print(f"Excel 보고서 생성 완료: {out_path}")


if __name__ == "__main__":
    main()
